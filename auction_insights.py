# -*- coding: utf-8 -*-
# This line is needed, so that Python can handle special chars from the German language.

#### DESCRIPTION ####
""" This is an autmated report aggregator, aggregating auction data from Google
Adwords. The file has many comments to help you guys adjust this file for other
purposes. The input are two directories for brand and non-brand data. The names
of these excel files should always have the following structure: <countrycode>
nonBrand Auktionsdatenbericht.xlsx

This program uses 4 modules:
    (1) pandas >    Delivered with Anaconda, no need to install into Python
    (2) openpyxl >  Outdated version delivered with Anaconda. Please use the
                    following code in the command line -> pip install openpyxl --upgrade
    (3) datetime >  Standard Python Library Module
    (4) os >        Standard Python Library Module

The code is quite dynamic and should be able to handle a flexible week number.
I.e. the output should still be sensical if the input are excel files with 2 or
8 weeks instead of the normal 5. Additionally the columns can be in different
orders, it identifies them on the basis of a dictionary.

Important notes:
    *   The <u> that can be found in front of strings (u'string') denotes it as
        a unicode string.

If there are ever questions feel free to ask them on my personal e-mail address
sebas.thuis@gmail.com.
"""

#### IMPORTS ####

import pandas as pd
from datetime import datetime
from os import listdir, path
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl import *

#### VARIABLES ####

# File location, relative to the working directory of the Python file.
# IMPORTANT: Always add <r> in front of a string if it is a filepath. Otherwise
# The </> will be interpreted wrong.
brand_directory = r'RawData/Brand/Last Week/'
nonbrand_directory = r'RawData/nonBrand/Last Week/'

# Suffix for the filenames of the raw input data. This program assumes that
# the file always starts with the country code. Additionally do not forget to put
# in a space before the word 'Brand'
brand_suffix = r' Brand Auktionsdatenbericht.xlsx'
nonbrand_suffix = r' nonBrand Auktionsdatenbericht.xlsx'

# Output filename prefix
brand_output_prefix = 'Brand_ReportBook'
nonbrand_output_prefix = 'nonBrand_ReportBook'

# List of countries, used for naming sheets and creating a list of filenames.
countries = ["AT", "AU", "BEfr", "BEnl", "CAen", "CAfr", "CHde", "CHfr", "CL", "CZ", "DE", "DK",
    "EFR", "ES", "ESE", "FI", "FR", "HU", "IE", "NL", "NO", "NZ", "PL", "RU", "SE", "SK", "UK",
    "US", "ZA"]

# This determines the order in which the devices will be shown in excel. If another device
# gets added, please add it to this list. Additionally it is used for cell labelling. You can
# add another device and it will add this to the column.
devices = [u'Computer', u'Tablet', u'Mobile']

# A dictionary that helps rename the columns. The columns we don't use are named 1,2,3 and 4 for easy referencing.
columns_dictionary = {u'Woche': u'week', u'Gerät': u'device', u'Domain der angezeigten URL': u'website',
    u'Anteil an möglichen Impr.': u'share', u'Durchschn. Position': u'rank_google', u'Überschneidungsrate': u'1',
    u'Rate der Position oberhalb': u'2', u'Rate für obere Positionen': u'3',u'Anteil an möglichen Impressionen gegenüber Mitbewerber': u'4'}

#### DEFINING FUNCTIONS ####

def complete_prompt():
    print ""
    print "----COMPLETED------------------------------------------"
    print ""
    print " All the files have been generated. The program has been"
    print " completed. You can close the command prompt."
    print ""
    print "-------------------------------------------------------"
    print ""

def status_prompt(output_filename):
    print ""
    print "----STATUS---------------------------------------------"
    print ""
    print " The %s file is generated." % output_filename
    print " You can find it in: %s" % path.dirname(path.abspath(__file__))
    print " Do not close the window, the program is still running"
    print ""
    print "-------------------------------------------------------"
    print ""

def error_choice(country, errortype):
    """ This is a simple function that presents the user with a choice if there
    is an error. It takes as input the country and the errortype, which can be
    'empty' or 'cannotopen'. """

    if errortype == 'cannotopen':
        print ""
        print "----ERROR----------------------------------------------"
        print ""
        print " The excel file of %s cannot be read, do you want to contine" % country
        print " to run the program? If you choose to continue, the country will"
        print " not be included in the final excel file. (y/n)?"
        print ""
        print "----INPUT REQUIRED-------------------------------------"
        print ""
    elif errortype == 'empty':
        print ""
        print "----ERROR----------------------------------------------"
        print ""
        print " The excel file of %s is empty, do you want to continue to run" % country
        print " the program? If you choose to continue, the country will not be"
        print " included in the final excel file. (y/n)?"
        print ""
        print "----INPUT REQUIRED-------------------------------------"
        print ""
    if raw_input(" >>> ") == "y":
        print ""
        print "-------------------------------------------------------"
        print ""
        print " Continuing the program."
        print ""
        print "-------------------------------------------------------"
        return True
    else:
        print "-------------------------------------------------------"
        print ""
        print " The program will shut down."
        print ""
        print "-------------------------------------------------------"
        quit()


def processing_data(dataframe):
    """This is the function that processes the dataframe that is generated
    by the module Pandas after opening an Excel file. It does the following

        (1) Renaming columns and rows
        (2) Drop irrelevant data
        (3) Reformatting European style numbers to an American format (needed for ranking)
        (4) Reformatting the structure of the table in a desirable format
        (5) Renaming device labels to a format that can be used by Python

    It returns a dictionary were each entry is a dictionary with the device name
    as a key. You can find the device names in the device list above. """

    # Renames the columns in the dataframe on the basis of the dictionary 'columns_dictionary'
    dataframe.columns = [columns_dictionary.get(x, x) for x in dataframe.columns]

    # Drop unnecessary data in columns, which are the columns 1, 2, 3 and 4 as determined in the 'columns_dictionary'
    dataframe = dataframe.drop(['1', '2', '3', '4'], 1)

    # Converts the dates to a string with their respective weeks.
    dataframe['week'] = dataframe['week'].astype('datetime64[ns]').dt.week # converts it to week numbers
    dataframe['week'] = 'CW ' + dataframe['week'].astype('str') # concatenates 'CW ' + weeknumber

    # Renames the device labels to a format that can be used by Python
    dataframe['device'] = dataframe['device'].apply(lambda x: x.replace(u'Tablets mit vollwertigem Internetbrowser',u'Tablet').replace(u'Mobilgeräte mit vollwertigem Internetbrowser', u'Mobile'))

    # Creates a column with the data that will eventually be used. (i.e. a concatenation of the sitename, rank and impression share)
    dataframe['content'] = (dataframe['website'] + ' \n' + dataframe['rank_google'] + ' | '+ dataframe['share'])

    # Creating a column with a ranking and then deleting everything below rank 10
    dataframe['rank_google'] = dataframe['rank_google'].apply(lambda x: float(unicode(x.replace(',','.')))) # Transforms the column into a number format that can be read
    dataframe['rank_top'] = dataframe.groupby(['week', 'device'])['rank_google'].rank(method='first').astype(int)
    dataframe = dataframe.sort_values('rank_top').groupby(['week', 'device']).head(n=10)

    # Reformatting the table. The aggfunc needs to be formulated in order to pivot. Since we are working with strings I have
    # defined an aggfunc that does some quick, but ultimately useless reformatting of the strings.
    dataframe = pd.pivot_table(dataframe, index = ['device', 'week'], columns='rank_top', values='content', aggfunc=lambda x: ' '.join(x))

    # Adding the prefix 'top' to the ranking
    dataframe.columns = ['top ' + str(i) for i in list(dataframe.columns.values)]

    # Creating dictionary where each entry is a subselection of the main dataframe
    data = {} # this is the dictionary
    for device in devices: # here each device gets its own entry.
        data[device] = dataframe.iloc[dataframe.index.get_level_values('device') == device]
        data[device] = data[device].reset_index(level=0, drop=True)

    return data

def writing_data(worksheet, dict_dataframes): # df stands for dataframe
    """ This function writes the dataframes in the right format to Excel. It uses
    the functionality of openpyxl version 2.4.0. Be wary, Anaconda does include
    openpyxl, but not the lastest version. Upgrade it via pip and the --upgrade
    command.

    The function needs two arguments, which are the worksheet and the dictionary
    of the data_frames. It will then the dataframes of each respective device
    onto one sheet.

    Summary of what the functions does:
        (1) Name the sheets
        (2) Write the data of the countries and their respective devices in that sheet
        (3) Rename some cells for readability

    It will return the workbook, worksheet and row_headers. These are later used
    as coordinates for formatting the tables."""

    # This list contains the row numbers on which the the headers (i.e. top 1, top 2, etc)
    # are placed. It initializes the by adding one element, with the value 1.
    # Since the first row has to be a header.

    row_headers = [1]

    # Openpyxl can interpret and write dataframes. They do that by iterating over the dataframe and writing it.
    # Quite slow, but at the moment this is the best solution if you want to style the excel file as well.
    # In the future pyexcelerate (not pyexcelerator) might be be a good replacement if they integrate pandas and

    # For loop that evaluates each dataframe and then each row in the dataframe and writes it to excel.
    for device in devices:
        for row in dataframe_to_rows(dict_dataframes[device], index=True, header=True):
            worksheet.append(row)
        row_headers.append(worksheet.max_row + 1)

    # Since the row_headers are supplemented with the last row of the last table + 1, the last value is dropped.
    row_headers = row_headers[:-1]

    # This adds the device description in the header (next to top 1). It uses the row_headers as coordinates.
    for n in range(3):
        worksheet.cell(row = row_headers[n - 1], column = 1).value = devices[n - 1] # This labels the respective subtabels with the relevant device names.

    return workbook, worksheet, row_headers

def styling_data(worksheet, row_headers):
    """ This function styles the tables in a format that is pleasing to the eye.
    Although the script looks quite complicated this is the most efficient and
    simple way to do this. The script loops over all cells that are present in
    a given worksheet and determines per cell through an if function how it should
    be formatted. The input is the worksheet and the row_headers which are row numbers.
    It will return just the worksheet.
    """

    # Making variables with the integers of the last rows and columns. These are defined for speed.
    last_row = worksheet.max_row
    last_column = worksheet.max_column

    # In order to speed up the script, I have opted to loop only once over every cell and format them.
    # I determine which cells should be formated on the basis of a if structure.

    # Looping over each cell.
    for row in worksheet.iter_rows(min_row = 1, min_col = 1, max_row = last_row, max_col = last_column):
        row_number = row[0].row # calculates the current row number, added here because it is used multiple times.

        # If statement that will format the first header in blue.
        if row_number == row_headers[0]:
            for cell in row:
                # This sets the width of the columns
                worksheet.column_dimensions[cell.column].width = 12
                cell.fill = PatternFill('solid', fgColor="4285F4")
                cell.font = Font(bold=True, color='ffffff')
                cell.border = Border(left=Side(border_style= 'thin', color='406fd6'),
                                     right=Side(border_style= 'thin', color='406fd6'),
                                     top=Side(border_style= 'thin', color='406fd6'),
                                     bottom=Side(border_style= 'thin', color='406fd6'),
                                     )

        # Elif statement that will format the second header in green.
        elif row_number == row_headers[1]:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor="0f9d58")
                cell.font = Font(bold=True, color='ffffff')
                cell.border = Border(left=Side(border_style= 'thin', color='169357'),
                                     right=Side(border_style= 'thin', color='169357'),
                                     top=Side(border_style= 'thin', color='169357'),
                                     bottom=Side(border_style= 'thin', color='169357'),
                                     )

        # Elif statement that formats the third header in red.
        elif row_number == row_headers[2]:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor="DB4437")
                cell.font = Font(bold=True, color='ffffff')
                cell.border = Border(left=Side(border_style= 'thin', color='bc2b2b'),
                                     right=Side(border_style= 'thin', color='bc2b2b'),
                                     top=Side(border_style= 'thin', color='bc2b2b'),
                                     bottom=Side(border_style= 'thin', color='bc2b2b'),
                                     )

        # Elif statement that formats the additional headers if more devices are added.
        elif row_number in row_headers:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor="DB4437")
                cell.font = Font(bold=True, color='ffffff')
                cell.border = Border(left=Side(border_style= 'thin', color='bc2b2b'),
                                     right=Side(border_style= 'thin', color='bc2b2b'),
                                     top=Side(border_style= 'thin', color='bc2b2b'),
                                     bottom=Side(border_style= 'thin', color='bc2b2b'),
                                     )

        # If the cell is not a header it must be either a column or the table content.
        else:
            for cell in row:
                # If statement that formats the column
                if cell.column == 'A':
                    # This sets the height of the table (except for the headers)
                    worksheet.row_dimensions[row_number].height = 45
                    cell.fill = PatternFill('solid', fgColor="f0f0f0")
                    cell.border = Border(left=Side(border_style='thin', color='e5e5e5'),
                                         right=Side(border_style='thin', color='e5e5e5'),
                                         top=Side(border_style='thin', color='e5e5e5'),
                                         bottom=Side(border_style='thin', color='e5e5e5')
                                         )

                # If the cell is not a header or column it is a content cell. This will be formatted here.
                else:
                    cell.alignment = Alignment(wrap_text=True)
                    cell.font = Font(size=8)
                    if cell.value is not None: # Sometimes cells are empty we still want to format them.
                        if u'Sie ' in cell.value:
                            cell.fill = PatternFill('solid', fgColor="dbdbdb")
                            cell.font = Font(bold =True, size=8)

    return worksheet

def generate_report(file_location, suffix_filename, output_filename): #
    """ This function actually generates all the input from the different excel
    files. It does this by using the functions that are defined above. As input
    the following arguments are needed.

        (1) The directory where the excel files are located. Relative to the
            location of the Python file.
        (2) The suffix of the filename, this means the string that comes after
            the country code for these excel files.
        (3) The output filename. Keep in mind that the the current week number
            will be added in the following format "_CW01"

    The function does the following:

        (1) It opens a workbook container.
        (2) Loops through each countrycode finds the respective file and processes
            this data.
        (3) Styles the worksheets.
        (3) Writes the entire workbook to an excel file. """

    # This opens a workbook
    workbook1 = Workbook()

    # This is a loop that goes through each country and processes the data.
    for country in countries:

        # This tests whether the excel file can be read. If it cannot be read, the
        # user will be prompted with a decision to continue the program.
        try:
            dataframe = pd.ExcelFile(file_location + country + suffix_filename, options={'encoding':'utf-8'})
        except:
            if error_choice(country, 'cannotopen') == True:
                continue

        # This parses the excel file. This is only needed to skip the first row.
        dataframe = dataframe.parse(skiprows = 1)

        # This tests whether the dataframe has any data. If it is empty it will
        # prompt the user with a choice.
        if dataframe.empty == True:
            if error_choice(country, 'empty') == True:
                continue

        # Retrieves the dictionary with the dataframes of each device.
        dictionary_devices = processing_data(dataframe)

        # Creates an dictionary for each worksheet, rather than making variables.
        worksheets = {}

        # Creates a excel worksheet for each country.
        worksheets[country] = workbook1.create_sheet(title=country)

        # Writes the data
        workbook, worksheets[country], row_headers = writing_data(worksheets[country], dictionary_devices)

        # Styles the data
        worksheets[country] = styling_data(worksheets[country], row_headers)

    # Removes the 'Sheet 1' that is automatically created in a workbook.
    workbook1.remove_sheet(workbook1.worksheets[0])

    # Unfortunately some modules have duplicate names, therefore I need to reimport datetime
    from datetime import datetime

    # Creating file
    output_filename = output_filename + '_CW%s.xlsx' % (datetime.now().strftime('%W'))
    workbook1.save(output_filename)

    # This generates a prompt giving the user feedback that the file has been generated.
    status_prompt(output_filename)


#### EXECUTION ####

generate_report(brand_directory, brand_suffix, brand_output_prefix)
generate_report(nonbrand_directory, nonbrand_suffix, nonbrand_output_prefix)
complete_prompt()
